import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

FILE_PATH = "storage/participants.xlsx"
SHEET_NAME = "Participants"


def normalize_phone(phone: str) -> str:
    phone = (phone or "").strip()
    phone = phone.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
    return phone


def normalize_name(full_name: str) -> str:
    return " ".join((full_name or "").strip().lower().split())


def style_worksheet(ws):
    ws.merge_cells("A1:G1")
    ws["A1"] = "TURNIR ISHTIROKCHILARI RO‘YXATI"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = [
        "Telegram ID",
        "Full Name",
        "Phone",
        "Username",
        "Tournament ID",
        "Tournament Name",
        "Registered At",
    ]

    thin = Side(style="thin", color="D9D9D9")

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_index)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    widths = {
        "A": 16,
        "B": 28,
        "C": 18,
        "D": 18,
        "E": 14,
        "F": 24,
        "G": 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = "A2:G2"


def ensure_parent_dirs():
    os.makedirs("storage", exist_ok=True)
    os.makedirs("exports", exist_ok=True)


def create_file_if_not_exists():
    ensure_parent_dirs()

    if not os.path.exists(FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        style_worksheet(ws)
        wb.save(FILE_PATH)
        wb.close()


def ensure_participants_file():
    create_file_if_not_exists()


def apply_data_row_style(ws, row_number):
    fill_color = "F8FBFF" if row_number % 2 == 0 else "EEF4FF"
    fill = PatternFill("solid", fgColor=fill_color)
    thin = Side(style="thin", color="E6E6E6")

    for col in range(1, 8):
        cell = ws.cell(row=row_number, column=col)
        cell.fill = fill
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cell.alignment = Alignment(vertical="center", horizontal="left")

    ws.row_dimensions[row_number].height = 20


def check_duplicate_in_tournament(ws, tournament_id, telegram_id, full_name, phone):
    norm_name = normalize_name(full_name)
    norm_phone = normalize_phone(phone)

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or row[1] is None:
            continue

        row_telegram_id = row[0]
        row_full_name = row[1]
        row_phone = row[2]
        row_tournament_id = row[4]

        if str(row_tournament_id) != str(tournament_id):
            continue

        if telegram_id is not None and str(row_telegram_id) == str(telegram_id):
            return False, "Bu Telegram ID bilan ishtirokchi allaqachon ro‘yxatdan o‘tgan."

        if norm_phone and normalize_phone(str(row_phone or "")) == norm_phone:
            return False, "Bu telefon raqam bilan ishtirokchi allaqachon ro‘yxatdan o‘tgan."

        if norm_name and normalize_name(str(row_full_name or "")) == norm_name:
            return False, "Bu ism-familiya bilan ishtirokchi allaqachon ro‘yxatdan o‘tgan."

    return True, "OK"


def add_participant(telegram_id, full_name, phone, username, tournament_id, tournament_name):
    create_file_if_not_exists()

    wb = load_workbook(FILE_PATH)
    ws = wb[SHEET_NAME]

    ok, msg = check_duplicate_in_tournament(
        ws=ws,
        tournament_id=tournament_id,
        telegram_id=telegram_id,
        full_name=full_name,
        phone=phone,
    )
    if not ok:
        wb.close()
        return False, msg

    registered_at = datetime.now().strftime("%d.%m.%Y %H:%M")
    new_row = ws.max_row + 1

    ws.cell(row=new_row, column=1, value=str(telegram_id) if telegram_id is not None else "")
    ws.cell(row=new_row, column=2, value=full_name)
    ws.cell(row=new_row, column=3, value=normalize_phone(phone))
    ws.cell(row=new_row, column=4, value=username or "")
    ws.cell(row=new_row, column=5, value=int(tournament_id))
    ws.cell(row=new_row, column=6, value=tournament_name)
    ws.cell(row=new_row, column=7, value=registered_at)

    apply_data_row_style(ws, new_row)

    wb.save(FILE_PATH)
    wb.close()
    return True, "Ishtirokchi Excel faylga saqlandi."


def get_participants_by_tournament(tournament_id):
    create_file_if_not_exists()

    wb = load_workbook(FILE_PATH)
    ws = wb[SHEET_NAME]

    participants = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or row[1] is None:
            continue

        if str(row[4]) == str(tournament_id):
            participants.append({
                "telegram_id": row[0],
                "full_name": row[1],
                "phone": row[2],
                "username": row[3],
                "tournament_id": row[4],
                "tournament_name": row[5],
                "registered_at": row[6],
            })

    wb.close()
    return participants


def _rewrite_rows(rows_to_keep):
    wb = load_workbook(FILE_PATH)
    ws = wb[SHEET_NAME]

    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)

    for row_data in rows_to_keep:
        new_row = ws.max_row + 1
        for col_index, value in enumerate(row_data, start=1):
            ws.cell(row=new_row, column=col_index, value=value)
        apply_data_row_style(ws, new_row)

    wb.save(FILE_PATH)
    wb.close()


def delete_participants_by_tournament(tournament_id):
    create_file_if_not_exists()

    wb = load_workbook(FILE_PATH)
    ws = wb[SHEET_NAME]

    rows_to_keep = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or row[1] is None:
            continue
        if str(row[4]) != str(tournament_id):
            rows_to_keep.append(row)

    wb.close()
    _rewrite_rows(rows_to_keep)


def delete_participant_by_tournament_and_user(tournament_id, telegram_id):
    create_file_if_not_exists()

    wb = load_workbook(FILE_PATH)
    ws = wb[SHEET_NAME]

    rows_to_keep = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or row[1] is None:
            continue

        row_telegram_id = row[0]
        row_tournament_id = row[4]

        if str(row_tournament_id) == str(tournament_id) and str(row_telegram_id) == str(telegram_id):
            continue

        rows_to_keep.append(row)

    wb.close()
    _rewrite_rows(rows_to_keep)


def export_tournament_participants(tournament, export_dir="exports"):
    ensure_parent_dirs()

    participants = get_participants_by_tournament(tournament["id"])
    file_name = f"participants_{tournament['id']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = str(Path(export_dir) / file_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"

    ws.merge_cells("A1:G1")
    ws["A1"] = f"{tournament['name']} | ISHTIROKCHILAR RO‘YXATI"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = [
        "№",
        "Full Name",
        "Phone",
        "Username",
        "Telegram ID",
        "Registered At",
        "Tournament",
    ]

    thin = Side(style="thin", color="D9D9D9")
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=i, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    widths = {
        "A": 8,
        "B": 28,
        "C": 18,
        "D": 18,
        "E": 16,
        "F": 22,
        "G": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for idx, participant in enumerate(participants, start=1):
        row = 2 + idx
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=participant["full_name"])
        ws.cell(row=row, column=3, value=participant["phone"])
        ws.cell(row=row, column=4, value=participant["username"])
        ws.cell(row=row, column=5, value=participant["telegram_id"])
        ws.cell(row=row, column=6, value=participant["registered_at"])
        ws.cell(row=row, column=7, value=participant["tournament_name"])
        apply_data_row_style(ws, row)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:G{max(2, ws.max_row)}"

    wb.save(file_path)
    wb.close()
    return file_path