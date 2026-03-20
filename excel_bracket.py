import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

TEMPLATES_DIR = Path("templates")
EXPORTS_DIR = Path("exports")
SHEET_NAME = "Сетка"

TEMPLATE_CANDIDATES = {
    (32, 8): ["32-8.xlsx"],
    (32, 16): ["32-16.xlsx"],
    (64, 16): ["64-16.xlsx"],
    (64, 32): ["64-32.xlsx"],
    (128, 32): ["128-32.xlsx"],
    (128, 64): ["128-64.xlsx", "128-64 (1).xlsx"],
    (256, 64): ["256-64.xlsx"],
    (256, 128): ["256-128.xlsx"],
}

# Qaysi format qaysi sariq blok zonasiga yoziladi
# M-zona: M/N ustunlar oralig'i
# W-zona: W/X ustunlar oralig'i
ZONE_BY_FORMAT = {
    (32, 8): ("W", "X"),
    (32, 16): ("M", "N"),
    (64, 16): ("W", "X"),
    (64, 32): ("M", "N"),
    (128, 32): ("W", "X"),
    (128, 64): ("M", "N"),
    (256, 64): ("W", "X"),
    (256, 128): ("M", "N"),
}


def _safe_name(value: str) -> str:
    value = (value or "").strip()
    value = re.sub(r'[\\/:*?"<>|]+', "_", value)
    return value or "turnir"


def _find_template_path(max_players: int, stage_size: int) -> Path:
    candidates = TEMPLATE_CANDIDATES.get((max_players, stage_size), [])
    for name in candidates:
        path = TEMPLATES_DIR / name
        if path.exists():
            return path

    pattern = f"{max_players}-{stage_size}"
    for path in TEMPLATES_DIR.glob("*.xlsx"):
        if pattern in path.name:
            return path

    raise FileNotFoundError(f"Shablon topilmadi: {max_players}-{stage_size}.xlsx")


def _participants_map(participants):
    result = {}
    if not participants:
        return result

    for p in participants:
        telegram_id = p.get("telegram_id")
        result[str(telegram_id)] = p

    return result


def _player_name_by_id(user_id, participants_map: dict) -> str:
    if user_id is None:
        return "BYE"

    player = participants_map.get(str(user_id))
    if not player:
        return f"ID {user_id}"

    full_name = (player.get("full_name") or "").strip()
    return full_name if full_name else f"ID {user_id}"


def _flatten_bracket_pairs(bracket_pairs):
    slots = []
    for left_player, right_player in bracket_pairs:
        slots.append(left_player)
        slots.append(right_player)
    return slots


def _slot_rows(stage_size: int):
    """
    Birinchi tur sariq slot qatorlari:
    1-juftlik -> 5,6
    2-juftlik -> 9,10
    3-juftlik -> 13,14
    ...
    """
    rows = []
    pair_count = stage_size // 2
    for i in range(pair_count):
        start_row = 5 + i * 4
        rows.append(start_row)
        rows.append(start_row + 1)
    return rows


def _choose_write_column(ws, left_col: str, right_col: str, row: int) -> str:
    """
    Sariq blok ichida haqiqiy yoziladigan katakni topadi.
    Ba'zi shablonlarda M yoziladi, ba'zilarida N.
    Xuddi shuningdek W yoki X.
    """
    left_cell = ws[f"{left_col}{row}"]
    right_cell = ws[f"{right_col}{row}"]

    def score(cell):
        s = 0
        if cell.border.left.style:
            s += 1
        if cell.border.right.style:
            s += 1
        if cell.border.top.style:
            s += 1
        if cell.border.bottom.style:
            s += 1
        return s

    left_score = score(left_cell)
    right_score = score(right_cell)

    # Qaysi katak "to‘liqroq" box bo‘lsa, o‘shanga yozamiz
    if left_score >= right_score:
        return left_col
    return right_col


def _build_target_cells(ws, max_players: int, stage_size: int):
    if (max_players, stage_size) not in ZONE_BY_FORMAT:
        raise ValueError(f"Qo‘llab bo‘lmaydigan kombinatsiya: {max_players}-{stage_size}")

    zone_left, zone_right = ZONE_BY_FORMAT[(max_players, stage_size)]
    rows = _slot_rows(stage_size)

    coords = []
    for row in rows:
        target_col = _choose_write_column(ws, zone_left, zone_right, row)
        coords.append(f"{target_col}{row}")

    return coords


def create_bracket_excel(tournament: dict, bracket_pairs, participants=None):
    if not tournament:
        raise ValueError("Turnir ma'lumoti bo‘sh")

    max_players = int(tournament["max_players"])
    stage_size = int(tournament["selected_stage"])

    template_path = _find_template_path(max_players, stage_size)

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    export_name = f"{_safe_name(tournament['name'])}_{max_players}-{stage_size}_{timestamp}.xlsx"
    export_path = EXPORTS_DIR / export_name

    shutil.copy2(template_path, export_path)

    wb = load_workbook(export_path)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise ValueError(f"'{SHEET_NAME}' sheet topilmadi")

    ws = wb[SHEET_NAME]

    participants_map = _participants_map(participants)
    ordered_slots = _flatten_bracket_pairs(bracket_pairs)
    target_cells = _build_target_cells(ws, max_players, stage_size)

    if len(ordered_slots) > len(target_cells):
        wb.close()
        raise ValueError(
            f"Kataklar yetarli emas. Slotlar: {len(ordered_slots)}, kataklar: {len(target_cells)}"
        )

    # Faqat sariq slot kataklarini tozalaymiz
    for coord in target_cells:
        ws[coord] = None

    # Ishtirokchilarni ketma-ket sariq slotlarga yozamiz
    for idx, user_id in enumerate(ordered_slots):
        coord = target_cells[idx]
        ws[coord] = _player_name_by_id(user_id, participants_map)

    wb.save(export_path)
    wb.close()

    return str(export_path)


# admin.py dagi universal chaqiruvlar bilan mos bo‘lishi uchun
generate_bracket_excel = create_bracket_excel
export_bracket_excel = create_bracket_excel
export_bracket_to_excel = create_bracket_excel
build_bracket_excel = create_bracket_excel